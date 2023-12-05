'''
This file contains the functions created in bib-detection.ipynb. 
Necessary functions to the API are stored here, they were simply copy and pasted from my Jupyter notebook.

If you wish to deploy the app from your own computer, note that the models will have to be trained in the 
bib-detector.ipynb notebook and the designated file paths must be changed prior to accessing these functions.

'''

# IMPORTS: -------------------------------------------------------------------------------------------------------------------

import pandas as pd
import numpy as np
from ultralytics import YOLO
from PIL import Image
from torchvision import transforms
import torch
import joblib

import pytesseract
from PIL import ImageEnhance, ImageFilter
from deskew import determine_skew
import imutils
import cv2

from openpyxl import Workbook
from openpyxl.drawing.image import Image as Img
import base64
from io import BytesIO

# ----------------------------------------------------------------------------------------------------------------------------

# MODEL INITIALIZATION: ------------------------------------------------------------------------------------------------------

'''
    MODEL OBJECTS:

        - model_bib     - bib detection model trained on custom dataset 
        - model_people  - pretrained model focusing on 'person' class
        - model_gender  - gender classification model trained on custom dataset
    
    #### NOTE: you will have to change the file path to the model below to wherever the best weights/model are saved on your computer. 
'''

# Load the YOLO bib detection model
model_bib = YOLO("best.pt") #### change file path here 

# Load the YOLO people model
model_people = YOLO('yolov8n.pt') # pretrained YOLOv8n model
# Make the only detection class be 'person'
model_people.classes = ['person']

# Load the Resnet50 gender classification model
model_gender = joblib.load('gender_classifier_resnet50.pkl')

# ----------------------------------------------------------------------------------------------------------------------------

# PREDICTION PIPELINES: ------------------------------------------------------------------------------------------------------

# predicition pipeline 
# (combines model_people and model_bib for a single image)
def prediction_pipeline(image):

    people_list = []
    bibs_list = []

    '''
        Takes a string image path and runs it through the person detection model & then the bib detection model.

        The lists above are where the cropped images of each will be saved (matched by index).

        This is so that when we use the gender classifer on the person & the number detector on the bib, both results will be able to be paired 
        efficiently in the final databases. 

    '''

    # Convert the input image path back to an actual image
    # (needed for person cropping and image detection in bib model)

    if isinstance(image, str): # if image is path 
        im = Image.open(image) # Open image using PIL
    else:
        im = image

    # If the image has an alpha channel (transparency), convert it to RGB
    if im.mode == 'RGBA':
        im = im.convert('RGB')
    
    # Run the image through people model
    model_people_result = model_people(im)

    for i, people in enumerate(model_people_result):

        people_boxes = people.boxes.xyxy.tolist() # extracts the bounding box coordniates from the results attributes
                                                  # setting the torch tensor to list will result in a list of lists,
                                                  # where each list within the list is a row of the tensor
        
        for b in people_boxes: # for every row of the tensor, it is the coordinates of a detected person box

            x_min, y_min, x_max, y_max = map(int, b) # extracting these coordinates from the rows 

            person_roi = im.crop((x_min, y_min, x_max, y_max)) # crops the region of interest (ROI) containing the person from the original image

            r = model_bib(person_roi)  # runs the cropped ROI through the bib model

            bib_box = r[0].boxes.xyxy.tolist()              # takes the model results and gets the bounding box coordinates

            if (bib_box != []): # making sure there IS a bib for the person
                x_min, y_min, x_max, y_max = map(int, bib_box[0])  # extracting these coordinates from the rows
                                                                    # no loop is necessary for the bib cropping b/c there should only be 1 bib per person 'b'

                bib_roi = person_roi.crop((x_min, y_min, x_max, y_max)) # crops the new region of interest (ROI) from the exisiting crop of the person
                                                                    # new ROI crop should be of the bib

                # Adding the cropped image of the person (person_roi) and the cropped image of the bib (im)
                people_list.append(person_roi)
                bibs_list.append(bib_roi)

    return people_list, bibs_list

# multiple predicition pipeline 
# (combines model_people and model_bib for a list of images)
def multiple_predicition_pipeline(image_paths):

    '''
        Takes multiple string image paths and runs them through the prediction_pipeline() function.

        Resulting lists are the same as what prediction_pipeline() returns, they just contain all of the people
        in the many images in the input list.
    '''

    total_people_list = []
    total_bib_list = []

    for image_path in image_paths:
        people_list, bibs_list = prediction_pipeline(image_path)
        total_people_list += people_list
        total_bib_list += bibs_list

    return total_people_list, total_bib_list

preprocess = transforms.Compose([
    transforms.Resize(256),
    transforms.CenterCrop(224),
    transforms.ToTensor(),
    transforms.Normalize(mean=[0.485, 0.456, 0.406], std=[0.229, 0.224, 0.225]),
])

# gender predicition pipeline 
# (combines model_people results and model_gender)
def gender_predictions(loaded_model, people_list):

    '''
        Just like with bibs_list and people_list (returned by the prediction pipeline functions), 
        gender_list will share matching indicies with the person image it predicts a gender on.
    '''

    gender_list = [] 
    for people in people_list:
       
       # Apply the transformations
        input_tensor = preprocess(people)
        input_batch = input_tensor.unsqueeze(0)  # Add a batch dimension

        # Make the prediction
        with torch.no_grad():
            output = loaded_model(input_batch)

        # Get the predicted label
        class_names = ['man', 'woman']
        _, predicted_idx = torch.max(output, 1)
        predicted_label = class_names[predicted_idx.item()]

        gender_list.append(predicted_label)

    return gender_list

# BIB IMAGE PRE-PROCESSING FUNCTIONS ------------------
def deskew(image, angle):
    '''
        This function helps adjusts the image according to the determined skew angle as apart of the image pre-processing for optimal
        OCR results. 
    '''
    non_zero_pixels = cv2.findNonZero(cv2.bitwise_not(image))
    center, wh, theta = cv2.minAreaRect(non_zero_pixels)

    root_mat = cv2.getRotationMatrix2D(center, angle, 1)
    rows, cols = image.shape
    rotated = cv2.warpAffine(image, root_mat, (cols, rows), flags=cv2.INTER_CUBIC)

    return rotated
def crop_dark_regions(image):
    '''
        This function removes the border of the image. A dark border can skew characterizations. 
    '''
    mask = np.zeros(image.shape, dtype=np.uint8)

    cnts = cv2.findContours(image, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    cnts = cnts[0] if len(cnts) == 2 else cnts[1]

    cv2.fillPoly(mask, cnts, [255,255,255])
    mask = 255 - mask
    result = cv2.bitwise_or(image, mask)

    return result
# -----------------------------------------------------

def number_detect(bibs_list, gender_list):
    '''

    Just like with bibs_list and people_list, nums_list will share matching indicies with the bib it gets the numbers for.

    If the detected characters are valid numbers, they will be added to the list. If they are not, 'None' will be passed as a placeholder

    '''

    nums_list = []
    for i, img in enumerate(bibs_list):

        img = bibs_list[i]

        img = np.array(img) # need cv image for next couple of steps
        img = imutils.resize(img, width=300) 

        # Deskewking image 
        img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        angle = determine_skew(img)
        img = deskew(img, angle)
        
        # Cropping dark edges from deskewing
        img = cv2.bilateralFilter(img, 11, 17, 17)
        img = crop_dark_regions(img)
        img = Image.fromarray(cv2.cvtColor(img, cv2.COLOR_BGR2RGB)) # converting back to PIL

        # Various photo editing
        img = img.filter(ImageFilter.GaussianBlur(radius=1)) # apply gaussian blur (noise removal)
        brightness = ImageEnhance.Brightness(img)
        img = brightness.enhance(1.5)  #increase brightness by 150%
        contrast = ImageEnhance.Contrast(img)
        img = contrast.enhance(2.0) # increase contrast by 200%

        pytesseract.pytesseract.tesseract_cmd = r'/usr/local/Cellar/tesseract/5.3.3/bin/tesseract' # Provide the path to the Tesseract executable
        
        nums = pytesseract.image_to_string(img, config='--psm 7 digits') # PSM 7 is used for a single line, digits used for numbers

        # Criteria to determine if the nums should be added 
        nums = nums.replace('.', '')
        nums = nums.replace('-','')
        nums = nums.strip()
        
        if nums == '':
            nums = None

        if nums not in nums_list: 
            nums_list.append(nums)
        else:
            del bibs_list[i]
            del gender_list[i]

        
    return(bibs_list, gender_list, nums_list)

# ----------------------------------------------------------------------------------------------------------------------------

# CREATE CSV RESULTS: --------------------------------------------------------------------------------------------------------

# Convert the images to base64 strings
def image_to_base64(pil_img):
    buffered = BytesIO()
    pil_img.save(buffered, format="JPEG")
    img_str = base64.b64encode(buffered.getvalue()).decode('utf-8')
    return img_str

def create_excel(gender_list, bibs_list, nums_list, email):
    wb = Workbook()
    man_sheet = wb.active
    man_sheet.title = "Man"
    woman_sheet = wb.create_sheet("Woman")

    man_sheet.append(['Bib Images', 'Recognized Bib Numbers'])
    woman_sheet.append(['Bib Images', 'Recognized Bib Numbers'])

    man_index = 0
    woman_index = 0

    for i, img in enumerate(bibs_list):
        img_str = image_to_base64(img)
        img_data = base64.b64decode(img_str)
        img = BytesIO(img_data)
        img_pil = Image.open(img)

        if gender_list[i] == 'man':
            man_sheet.add_image(Img(img_pil), f'A{man_index+2}')
            man_sheet.row_dimensions[man_index+2].height = img_pil.height
            man_sheet.column_dimensions['A'].width = img_pil.width

        elif gender_list[i] == 'woman':
            woman_sheet.add_image(Img(img_pil), f'A{woman_index+2}')
            woman_sheet.row_dimensions[woman_index+2].height = img_pil.height
            woman_sheet.column_dimensions['A'].width = img_pil.width

        if gender_list[i] == 'man':
            man_sheet.cell(row=man_index+2, column=3, value=nums_list[man_index])
            man_index+=1

        elif gender_list[i] == 'woman':
            woman_sheet.cell(row=woman_index+2, column=3, value=nums_list[woman_index])
            woman_index+=1

    wb.save(f"results/{email}_results.xlsx")

# ----------------------------------------------------------------------------------------------------------------------------

# MASTER PIPELINE: -----------------------------------------------------------------------------------------------------------

# Converting PIL objects to base64 strings and remove duplicates -
# doing this so that if the same person is detected in multiple frames, only one instance is kept.
# Taking only every 4th frame should help prevent SOME of this, but there will still be duplicates.
def remove_duplicates(people_list, bibs_list):
    unique_people = []
    unique_bibs = []
    seen = set()
    for i in range(len(people_list)):
        img_str = image_to_base64(people_list[i])
        if img_str not in seen:
            unique_people.append(people_list[i])
            unique_bibs.append(bibs_list[i])
            seen.add(img_str)
    return unique_people, unique_bibs

# MASTER PIPELINE 

def master(pil_images, email):
    '''
        Takes a list of PIL images and runs all of the prediction models of them.
        Results in a spreadsheet of detected bibs. 
    '''

    people_list, bibs_list = multiple_predicition_pipeline(pil_images) # object detection for the people and bibs in the frames
    people_list, bibs_list = remove_duplicates(people_list, bibs_list) # removing duplicates 

    gender_list = gender_predictions(model_gender, people_list) # classification for gender of people

    bibs_list, gender_list, nums_list = number_detect(bibs_list, gender_list) # OCR on images on bibs

    create_excel(gender_list, bibs_list, nums_list, email) # creating excel of all this data 

# ----------------------------------------------------------------------------------------------------------------------------